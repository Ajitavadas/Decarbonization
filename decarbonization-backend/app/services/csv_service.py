import csv
from io import StringIO, BytesIO
from typing import List, Dict, Tuple
from datetime import datetime
from app.services.ai_classifier_service import ai_classifier


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional XLSX support
    load_workbook = None

class CSVParsingService:
    """Service for parsing and validating CSV files"""
    
    REQUIRED_COLUMNS = {
        'description': str,
        'transaction_date': 'datetime',
        'scope': int,
        'category': str,
        'activity_value': float,
        'activity_unit': str,
        'emission_factor_value': float
    }
    
    OPTIONAL_COLUMNS = {
        'supplier_name': str,
        'project_id': str,
        'notes': str
    }
    
    VALID_SCOPES = [1, 2, 3]
    
    @staticmethod
    def parse_csv(file_content: bytes, encoding: str = 'utf-8', validate: bool = True) -> Tuple[List[Dict], List[Dict]]:
        """
        Parse CSV file.
        If validate=True, filters validity.
        If validate=False, returns all rows (as dicts) and empty errors (unless parse error).
        """
        # ... decode logic ...
        try:
            text = file_content.decode(encoding)
        except UnicodeDecodeError:
            try:
                text = file_content.decode('iso-8859-1')
            except Exception:
                return [], [{"row": 0, "error": "Cannot decode file (unsupported encoding)"}]
        
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        try:
            csv_reader = csv.DictReader(StringIO(text))
        except Exception as e:
            return [], [{"row": 0, "error": f"Invalid CSV format: {str(e)}"}]
        
        if not csv_reader.fieldnames:
            return [], [{"row": 0, "error": "Empty CSV file or no header row"}]
            
        rows = []
        error_rows = []
        
        for row_num, row in enumerate(csv_reader, start=2):
            if all(not cell.strip() if isinstance(cell, str) else not cell for cell in row.values()):
                continue
            
            if validate:
                is_valid, error_msg = CSVParsingService.validate_row(row, row_num)
                if is_valid:
                    row["_row_num"] = row_num
                    rows.append(row)
                else:
                    error_rows.append({"row": row_num, "error": error_msg})
            else:
                row["_row_num"] = row_num
                rows.append(row)
                
        return rows, error_rows
    
    @staticmethod
    def parse_xlsx(file_content: bytes, validate: bool = True) -> Tuple[List[Dict], List[Dict]]:
        """
        Parse XLSX file with error handling.
        """
        if load_workbook is None:
            return [], [{"row": 0, "error": "XLSX parsing not available (missing openpyxl dependency)"}]

        try:
            wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
            sheet = wb.active
        except Exception as e:
            return [], [{"row": 0, "error": f"Invalid XLSX file: {str(e)}"}]

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return [], [{"row": 0, "error": "Empty XLSX file"}]

        headers = [str(h).strip() if h is not None else "" for h in headers]
        header_index = {h: idx for idx, h in enumerate(headers)}

        rows = []
        error_rows = []

        row_num = 1  # header row
        for excel_row in rows_iter:
            row_num += 1
            row_dict: Dict[str, str] = {}

            # If not validating, we want ALL columns, not just required items?
            # Actually, `parse_csv` uses `csv.DictReader` which blindly takes headers.
            # Here we are manually constructing generic `row_dict`.
            # If we want generic rows, we should capture all columns in headers.
            
            if not validate:
                # Capture all columns corresponding to headers
                for h, idx in header_index.items():
                    if idx < len(excel_row):
                        val = excel_row[idx]
                        row_dict[h] = "" if val is None else str(val)
            else:
                # Capture specific known columns
                for col_name in list(CSVParsingService.REQUIRED_COLUMNS.keys()) + list(CSVParsingService.OPTIONAL_COLUMNS.keys()):
                    if col_name in header_index:
                        val = excel_row[header_index[col_name]]
                        row_dict[col_name] = "" if val is None else str(val)

            # Skip completely empty rows
            if all(not v.strip() for v in row_dict.values()):
                continue

            if validate:
                is_valid, error_msg = CSVParsingService.validate_row(row_dict, row_num)
                if is_valid:
                    row_dict["_row_num"] = row_num
                    rows.append(row_dict)
                else:
                    error_rows.append({"row": row_num, "error": error_msg})
            else:
                row_dict["_row_num"] = row_num
                rows.append(row_dict)

        return rows, error_rows

    @staticmethod
    def validate_row(row: Dict, row_num: int) -> Tuple[bool, str]:
        """
        Validate single row
        
        Args:
            row: Dictionary representing a CSV row
            row_num: Row number (for error reporting)
            
        Returns:
            (is_valid, error_message)
        """
        
        # Check for missing required fields
        for col in CSVParsingService.REQUIRED_COLUMNS.keys():
            if col not in row:
                return False, f"Missing column: {col}"
            
            if not row[col] or (isinstance(row[col], str) and not row[col].strip()):
                return False, f"Missing required field: {col}"
        
        # Validate description
        if len(row['description'].strip()) > 500:
            return False, "Description exceeds 500 characters"
        
        # Validate scope (1-3)
        try:
            scope = int(row['scope'].strip())
            if scope not in CSVParsingService.VALID_SCOPES:
                return False, f"Scope must be 1, 2, or 3. Got: {scope}"
        except ValueError:
            return False, f"Scope must be integer. Got: {row['scope']}"
        
        # Validate category
        if len(row['category'].strip()) > 100:
            return False, "Category exceeds 100 characters"
        
        # Validate activity_value (positive float)
        try:
            activity_val = float(row['activity_value'].strip())
            if activity_val <= 0:
                return False, f"Activity value must be positive. Got: {activity_val}"
        except ValueError:
            return False, f"Activity value must be number. Got: {row['activity_value']}"
        
        # Validate activity_unit
        if len(row['activity_unit'].strip()) > 50:
            return False, "Activity unit exceeds 50 characters"
        
        # Validate emission_factor_value (positive float)
        try:
            factor_val = float(row['emission_factor_value'].strip())
            if factor_val <= 0:
                return False, f"Emission factor must be positive. Got: {factor_val}"
        except ValueError:
            return False, f"Emission factor must be number. Got: {row['emission_factor_value']}"
        
        # Validate transaction_date
        try:
            date_str = row['transaction_date'].strip()
            # Try ISO format first
            if 'T' in date_str or ' ' in date_str:
                datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            else:
                # Try date-only format
                datetime.strptime(date_str, '%Y-%m-%d')
        except (ValueError, KeyError):
            return False, f"Invalid date format. Use ISO format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS). Got: {row['transaction_date']}"

        # Optional fields length validation
        supplier = row.get('supplier_name') or ""
        if supplier and len(supplier.strip()) > 255:
            return False, "Supplier name exceeds 255 characters"

        project_id = row.get('project_id') or ""
        if project_id and len(project_id.strip()) > 100:
            return False, "Project ID exceeds 100 characters"

        notes = row.get('notes') or ""
        if notes and len(notes.strip()) > 1000:
            return False, "Notes field exceeds 1000 characters"

        # Basic spreadsheet formula injection guard on key text fields
        for field_name in ["description", "category", "supplier_name", "project_id", "notes"]:
            val = (row.get(field_name) or "").lstrip()
            if val and val[0] in ("=", "+", "-", "@"):
                return False, f"Potential formula injection detected in field '{field_name}'"

        return True, ""
    
    @staticmethod
    def calculate_co2e(activity_value: float, emission_factor: float) -> Tuple[float, float]:
        """
        Calculate CO2e emissions
        
        Formula: CO2e = activity_value × emission_factor
        
        Args:
            activity_value: Quantity of activity
            emission_factor: CO2e per unit of activity
            
        Returns:
            (co2e_kg, co2e_tonnes)
        """
        co2e_kg = activity_value * emission_factor
        co2e_tonnes = co2e_kg / 1000
        
        return round(co2e_kg, 3), round(co2e_tonnes, 6)

    @staticmethod
    async def classify_and_enhance_rows(
        rows: List[Dict]
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        Enhance rows with AI scope classification
        
        Returns:
            (enhanced_rows, ai_errors)
        """
        enhanced_rows = []
        ai_errors = []
        
        for row in rows:
            try:
                # Call AI classifier
                scope, confidence, needs_review = await ai_classifier.classify_transaction(
                    description=row.get('description', ''),
                    category=row.get('category', ''),
                    supplier_name=row.get('supplier_name')
                )
                
                # Update row with AI predictions
                row['ai_scope_prediction'] = scope
                row['ai_confidence_score'] = confidence
                row['ai_needs_review'] = needs_review
                
                enhanced_rows.append(row)
                
            except Exception as e:
                ai_errors.append({
                    "row": row.get('_row_num', '?'),
                    "error": f"AI classification failed: {str(e)}"
                })
                # Set defaults on failure
                row['ai_scope_prediction'] = None
                row['ai_confidence_score'] = 0.0
                row['ai_needs_review'] = True
                enhanced_rows.append(row)
        
        return enhanced_rows, ai_errors
    